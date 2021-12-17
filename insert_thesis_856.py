from tkinter import filedialog
from pathlib import Path
import datetime
from openpyxl import load_workbook
from pymarc import MARCReader, Record, Field


def spreadsheet_lookup_dict(lookup_file_path=None):
    tmp_dict = {}
    return_dict = None

    wb = load_workbook(lookup_file_path)
    ws = wb.active

    for row in range(1, ws.max_row + 1):
        tmp_dict[str(ws[row][0].value)] = str(ws[row][1].value)
        print(row)

    if len(tmp_dict) > 0:
        return_dict = tmp_dict

    print(f"Distinct Records in Look up Dict: {len(return_dict)}")

    return return_dict


def process_marc(marc_file_path=None, output_file_path=None, lookup_file_path=None):
    # Load spreadsheet into dict for faster searching.
    spreadsheet_dict = spreadsheet_lookup_dict(
        lookup_file_path=lookup_file_path
    )

    # Read the MARC File (mf)
    with open(marc_file_path, 'rb') as mf:
        reader = MARCReader(mf)

        # Loop through records
        for record in reader:

            f_001 = record['001'].data
            print(f_001)

            # Get 856s to loop through to update/create
            if record.get_fields('856'):
                for f_856 in record.get_fields('856'):
                    # u_856 = f_856['u'].data
                    u_856 = f_856['u']
                    print(u_856)

                    # Check to make sure it is the handle.
                    if u_856:
                        if 'hdl.handle.net' in u_856:
                            # Remove field if matched:
                            record.remove_field(f_856)

            # Default no new 856 in case no result found.
            new_856 = None
            # Look up link in dict to get link.
            if f_001 in spreadsheet_dict.keys():
                new_856 = spreadsheet_dict[f_001]

            # If result found create the 856 with found link.
            if new_856:
                record.add_field(
                    Field(
                        tag='856',
                        indicators=['4', '1'],
                        subfields=[
                            'u', new_856,
                            'z', 'Link to OAKTrust copy'
                        ]
                    )
                )

            # Write changes to new output file.
            with open(output_file_path, 'ab') as out:
                out.write(record.as_marc())
                out.close()


def main():
    # Get MARC file to update.
    mrc = filedialog.askopenfile(title="Select input MRC file")

    if mrc == "" or mrc is None:
        print("User canceled operations.")
        exit()

    marc_file_path = Path(mrc.name)

    # Establish output MARC file.
    # Set file name equal to now YYYY-MM-DD_HH-MM
    dt_now = datetime.datetime.now()
    output_file_path = Path.joinpath(
        marc_file_path.parent.absolute(),
        f"output__{dt_now.year}-{dt_now.month}-{dt_now.day}_{dt_now.hour}-{dt_now.minute}.mrc"
    )

    # Get spreadsheet with identifier and value to insert.
    lookup = filedialog.askopenfile(title="Select the Lookup Spreadsheet")

    if lookup == "" or lookup is None:
        print("User canceled operations.")
        exit()

    lookup_file_path = Path(lookup.name)

    # Run the processing.
    process_marc(
        marc_file_path=marc_file_path,
        output_file_path=output_file_path,
        lookup_file_path=lookup_file_path
    )


if __name__ == "__main__":
    main()
